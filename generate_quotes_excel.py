"""
Generate quotes.xlsx — pre-filled with Socrates quotes and story-driven captions.
Captions use a 4-part story framework: Hook → Story → Wisdom → CTA.

Each caption is built from audience-specific templates that weave the quote
into a relatable narrative. This replaces flat advice-style captions with
compelling micro-stories that stop the scroll.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Story-Driven Caption Templates ───────────────────────────────────────────
# Each audience has 3 unique story scenarios that get cycled across quotes.
# Templates use {quote} placeholder — filled at build time.

HOOKS = {
    "procrastinator": [
        "Be honest — how many times have you told yourself 'I'll start tomorrow' this month?",
        "There's a version of you that already did the thing. Why isn't that you?",
        "You just scrolled past this post. Then paused. That pause means something.",
        "The gap between who you are and who you want to be isn't ability. It's start.",
    ],
    "doomscroller": [
        "You've already seen 40 posts today. Remember any of them?",
        "Put the phone down for a second. Feel that silence? It's not empty. It's waiting.",
        "Scroll down. Scroll up. Scroll down. Now tell me — what did you actually learn?",
        "You're feeding your mind junk food. No shame. But recognise it for what it is.",
    ],
    "stuck": [
        "You're not stuck. You're standing at a door you're afraid to open.",
        "The feeling of being stuck isn't a stop sign. It's a compass.",
        "You've been waiting for clarity before you move. What if movement creates clarity?",
        "You know that quiet voice saying 'something needs to change'? Stop ignoring it.",
    ],
    "lazy": [
        "Your comfort zone isn't comfortable anymore. It's just familiar.",
        "You don't need motivation. You need 5 minutes of doing the thing.",
        "The bed feels warm. The couch is safe. But you know where warm and safe leads — nowhere.",
        "You're waiting for a spark of motivation. What if action creates the spark, not the other way around?",
    ],
    "quitter": [
        "You're one bad day away from quitting. I've been there. Don't.",
        "The version of you that started this would be so disappointed in the version that's about to quit.",
        "You didn't come this far to only come this far. Cliché? Maybe. True? Absolutely.",
        "The resistance you're feeling is proportional to how much this matters.",
    ],
    "lost": [
        "You don't need a map. You need to start walking.",
        "Not knowing where you're going is terrifying. It's also the only way to discover something real.",
        "You've been asking everyone else what you should do. When was the last time you asked yourself?",
        "The person you're meant to become is hiding on the other side of a question you're afraid to ask.",
    ],
    "overwhelmed": [
        "You're trying to hold everything at once. Some things need to drop.",
        "Breathe. You don't have to fix everything today. Just this moment.",
        "Your to-do list is a monster you're feeding. When do you put the fork down?",
        "You can't pour from an empty cup. When did yours last get refilled?",
    ],
}

STORIES = {
    "procrastinator": [
        "I spent two years waiting for the 'right time' to write. Better desk. More free time. The right tools. Two years of nothing. Then I wrote 500 terrible words on a Tuesday night. They were awful. But they were done. That's when I understood — starting imperfectly beats perfect planning.",
        "My mentor once told me: 'You don't rise to the level of your goals. You fall to the level of your systems.' I had goals written everywhere. Zero systems. Once I stopped romanticising the outcome and started obsessing over the first step, everything shifted.",
    ],
    "doomscroller": [
        "I tracked my screen time last week. 47 hours. 47 hours of watching other people live while I sat on my couch. That's a full work week. Gone. Seneca warned us 2,000 years ago: 'It's not that we have a short time, but that we waste much of it.' He was talking to people who didn't even have smartphones.",
        "The algorithm wants you angry. Angry people scroll longer. They comment more. They come back. I fell for it for months before I realised — I wasn't consuming content. Content was consuming me.",
    ],
    "stuck": [
        "I spent a year in a job I hated because I was afraid of the unknown. Every morning felt like walking through mud. Then I lost that job and realised — the unknown wasn't scary. It was freedom. The prison was my own fear of change.",
        "I remember the exact moment I stopped feeling stuck. I was staring at the ceiling at 3am, running the same questions in circles. Then I wrote one question down: 'What's the smallest possible step?' Answering that broke the loop.",
    ],
    "lazy": [
        "I used to think I needed motivation to act. So I waited. And waited. And waited for the spark. Then I heard a quote that changed everything: 'Action isn't the result of motivation. Motivation is the result of action.' So I did 5 pushups. Just 5. The next day I did 10. A year later I'd run a marathon.",
        "I was the king of 'I'll do it later'. Later never came. Then I realised laziness wasn't my problem — overwhelm was. The task felt too big. So I shrank it until it felt stupidly small. Read one page. Write one sentence. Walk for 60 seconds. Small beats nothing.",
    ],
    "quitter": [
        "I almost quit my business 3 months in. Revenue was zero. Doubt was everywhere. I had a resignation letter drafted. Then I re-read why I started — and realised I hadn't even given it a real shot. I was quitting before I'd earned the right to quit. That was 4 years ago.",
        "The darkest moment of my life was sitting in my car at 11pm, asking myself if I should just give up on everything. I didn't. Not because I was strong. Because I remembered that quitting wouldn't end the pain — it would just make the past pain pointless.",
    ],
    "lost": [
        "I spent my 20s trying to be what everyone expected. Good job. Right path. Safe choices. I was miserable. Then I read that Socrates chose to die rather than stop questioning. That level of conviction made me realise — I'd never questioned anything that mattered.",
        "When I didn't know what to do with my life, I asked everyone for advice. Parents. Friends. Strangers on the internet. None of them knew. The answer was inside me the whole time — I just didn't know how to listen yet.",
    ],
    "overwhelmed": [
        "I used to say yes to everything. Every meeting. Every favour. Every opportunity. I thought this was how success worked. I ended up broke, exhausted, and resentful. Then I learned one word: no. It felt selfish at first. Now I know it's survival.",
        "I remember standing in my kitchen crying because I couldn't decide what to cook for dinner. Not because I was sad — because my brain was so full from 14 hours of 'urgent' tasks that choosing a vegetable felt impossible. That's when I learned burnout isn't weakness. It's your soul asking for space.",
    ],
}

LESSONS = {
    "procrastinator": [
        "{quote} Socrates understood that we don't build the new by fighting the old — we build by building. Every philosopher starts with one question. Every creator starts with one bad draft. Every change starts with one imperfect step.",
        "{quote} The ancient Greeks had a word for this: 'energeia' — the actualising of potential through action. Not through planning. Not through waiting. Through doing. Right now, imperfectly, without knowing how it ends.",
    ],
    "doomscroller": [
        "{quote} Socrates didn't have Instagram. But he understood something crucial — the mind takes the shape of whatever it dwells on. Feed it outrage, you become outraged. Feed it wisdom, you become wise. Your feed is a garden. What are you planting?",
        "{quote} The Greeks called it 'akrasia' — knowing what's good for you but doing the opposite. Socrates saw it 2,400 years before the dopamine loop was named. The answer isn't willpower. It's awareness. Notice what you're feeding your mind. Then choose differently.",
    ],
    "stuck": [
        "{quote} Socrates taught that not-knowing isn't failure — it's the only honest starting point. Stuck means you've arrived at a question you haven't answered yet. That's not a dead end. That's the threshold of growth.",
        "{quote} The Oracle told Socrates he was the wisest man because he alone admitted he knew nothing. Your stuck-ness is a form of intellectual honesty. You've stopped pretending. Now the real work can begin.",
    ],
    "lazy": [
        "{quote} The stoics called this 'prohairesis' — the power of choice in every moment. You can't control your lack of motivation. But you can control the next 60 seconds. One small choice. Then another. That's how philosophy is built. That's how a life is built.",
        "{quote} Socrates lived this. He didn't write books. He didn't build an empire. He showed up every day in the marketplace and had one conversation. Day after day. That consistency — not genius — is what changed the world.",
    ],
    "quitter": [
        "{quote} Socrates was given a choice: stop philosophising or die. He chose death. Not because he was reckless — because he knew that quitting his mission would be a worse death than any hemlock could deliver. What mission are you about to abandon?",
        "{quote} The Stoics believed that every obstacle is actually fuel for progress — they called it 'the obstacle is the way'. What feels like a wall right now might be the very thing that forges you. Don't quit before the fire does its work.",
    ],
    "lost": [
        "{quote} Socrates spent his entire life on this one question. Not 'What should I do?' but 'Who am I?' He believed that knowing yourself was the foundation of everything else. Not your job title. Not your relationship status. You. The real you.",
        "{quote} The unexamined life wasn't worth living to Socrates not because he was dramatic — but because he knew that a life run on autopilot is a life someone else is living. Your confusion isn't a bug. It's the first honest thought you've had in years.",
    ],
    "overwhelmed": [
        "{quote} The Stoics believed that peace isn't found by doing more — it's found by wanting less. Socrates owned nothing and changed everything. What if the answer to your overwhelm isn't a better system — but surrender to what actually matters?",
        "{quote} Socrates knew that knowledge — real understanding of what matters — is the only lasting wealth. Everything else is borrowed. Your deadlines, your obligations, your 'urgent' emails — most of it won't matter in a week. Strip down to what matters.",
    ],
}

CTAS = {
    "procrastinator": [
        "What's one thing you've been putting off? Say it in the comments. Make it real. 👇",
        "What's your 'I'll start tomorrow' story? Drop it below. Let's kill it together. 🎯",
    ],
    "doomscroller": [
        "Right now — name ONE thing you'd rather be doing than scrolling. Put it in the comments. 👇",
        "What would you do with your phone time this week if you reclaimed just 30 minutes a day? 📱",
    ],
    "stuck": [
        "What's one small step you can take today? Just one. Write it below. 👇",
        "What's the question you're most afraid to answer about your life right now? 💭",
    ],
    "lazy": [
        "What's the smallest thing you can do right now? Not tomorrow. Right now. Tell me. ⚡",
        "Name ONE thing you've been avoiding that would take less than 5 minutes. Do it after this comment. 👇",
    ],
    "quitter": [
        "What made you start? Write it down before you talk yourself out of it again. 🔥",
        "What's the voice telling you right now? Drop it below. We'll talk you through it. 💪",
    ],
    "lost": [
        "When did you last ask yourself 'what do I actually want?' — and wait for the real answer? 🧭",
        "What's one belief about yourself that might be wrong? Just one. Start questioning. 🌊",
    ],
    "overwhelmed": [
        "What's ONE thing you can let go of today? Name it. Release it. 🌿",
        "What would 'enough' look like for you today? Define it. Protect it. 🎯",
    ],
}


def _build_caption(quote: str, audience: str, index: int) -> str:
    """
    Build a story-driven caption from audience-specific templates.
    Each caption = Hook → Story → Wisdom (quote) → CTA
    The index cycles through available hooks/stories/lessons/ctas for variety.
    """
    hooks = HOOKS.get(audience, HOOKS["stuck"])
    stories = STORIES.get(audience, STORIES["stuck"])
    lessons = LESSONS.get(audience, LESSONS["stuck"])
    ctas = CTAS.get(audience, CTAS["stuck"])

    hook = hooks[index % len(hooks)]
    story = stories[index % len(stories)]
    lesson = lessons[index % len(lessons)].format(quote=quote)
    cta = ctas[index % len(ctas)]

    return f"{hook}\n\n{story}\n\n{lesson}\n\n{cta}"


# ── Quotes: (quote, audience, status) ────────────────────────────────────────
# Captions are generated by _build_caption — no hardcoded strings needed.

QUOTES = [
    # PROCRASTINATORS
    ("The secret of change is to focus all your energy not on fighting the old, but on building the new.", "procrastinator", "ready"),
    ("An unexamined life is not worth living.", "procrastinator", "ready"),
    ("The beginning of wisdom is the definition of terms.", "procrastinator", "ready"),
    ("He who is not contented with what he has, would not be contented with what he would like to have.", "procrastinator", "ready"),
    ("Wonder is the beginning of wisdom.", "procrastinator", "ready"),
    ("I know that I know nothing.", "procrastinator", "ready"),
    ("To find yourself, think for yourself.", "procrastinator", "ready"),
    ("The greatest way to live with honor in this world is to be what we pretend to be.", "procrastinator", "ready"),
    ("Employ your time in improving yourself by other men's writings.", "procrastinator", "ready"),
    ("Every action has its pleasures and its price.", "procrastinator", "ready"),
    ("False words are not only evil in themselves, but they infect the soul with evil.", "procrastinator", "ready"),
    ("Contentment is natural wealth, luxury is artificial poverty.", "procrastinator", "ready"),
    ("The only good is knowledge and the only evil is ignorance.", "procrastinator", "ready"),

    # DOOMSCROLLERS
    ("Beware the barrenness of a busy life.", "doomscroller", "ready"),
    ("Strong minds discuss ideas, average minds discuss events, weak minds discuss people.", "doomscroller", "ready"),
    ("Be slow to fall into friendship, but when you are in, continue firm and constant.", "doomscroller", "ready"),
    ("Not life, but good life, is to be chiefly valued.", "doomscroller", "ready"),
    ("The mind is everything. What you think, you become.", "doomscroller", "ready"),
    ("Know thyself.", "doomscroller", "ready"),
    ("From the deepest desires often come the deadliest hate.", "doomscroller", "ready"),
    ("He is richest who is content with the least.", "doomscroller", "ready"),
    ("The hour of departure has arrived, and we go our ways — I to die, and you to live. Which is better, only God knows.", "doomscroller", "ready"),

    # FEELING STUCK
    ("The only true wisdom is in knowing you know nothing.", "stuck", "ready"),
    ("Wonder is the beginning of wisdom.", "stuck", "ready"),
    ("To move the world, we must first move ourselves.", "stuck", "ready"),
    ("Let him who would move the world first move himself.", "stuck", "ready"),
    ("The unexamined life is not worth living.", "stuck", "ready"),
    ("I am not an Athenian or a Greek, but a citizen of the world.", "stuck", "ready"),
    ("The easiest and noblest way is not to be crushing others, but to be improving yourselves.", "stuck", "ready"),
    ("From the deepest desires come the clearest paths.", "stuck", "ready"),
    ("By all means marry. If you get a good wife, you'll become happy. If you get a bad one, you'll become a philosopher.", "stuck", "ready"),

    # LAZY / LOW ENERGY
    ("Do not do to others what angers you if done to you by others.", "lazy", "ready"),
    ("He who is not contented with what he has would not be contented with what he would like to have.", "lazy", "ready"),
    ("The secret of change is to focus all your energy not on fighting the old, but on building the new.", "lazy", "ready"),
    ("No man has the right to be an amateur in the matter of physical training. It is a shame for a man to grow old without seeing the beauty and strength of which his body is capable.", "lazy", "ready"),
    ("Be as you wish to seem.", "lazy", "ready"),
    ("We cannot live better than in seeking to become better.", "lazy", "ready"),
    ("One who is injured ought not to return the injury, for on no account can it be right to do an injustice.", "lazy", "ready"),
    ("Think not those faithful who praise all thy words and actions, but those who kindly reprove thy faults.", "lazy", "ready"),

    # ABOUT TO QUIT
    ("The nearest way to glory is to strive to be what you wish to be thought to be.", "quitter", "ready"),
    ("Understanding a question is half an answer.", "quitter", "ready"),
    ("Sometimes you put walls up not to keep people out, but to see who cares enough to break them down.", "quitter", "ready"),
    ("The greatest blessing granted to mankind come by way of madness, which is a divine gift.", "quitter", "ready"),
    ("Every action has its pleasures and its price.", "quitter", "ready"),
    ("Death may be the greatest of all human blessings.", "quitter", "ready"),
    ("My plainness of speech makes them hate me.", "quitter", "ready"),
    ("From the deepest desires often come the deadliest hate.", "quitter", "ready"),

    # KNOW YOURSELF / LOST
    ("To know thyself is the beginning of wisdom.", "lost", "ready"),
    ("The unexamined life is not worth living.", "lost", "ready"),
    ("I cannot teach anybody anything. I can only make them think.", "lost", "ready"),
    ("An honest man is always a child.", "lost", "ready"),
    ("The greatest blessing granted to mankind come by way of madness.", "lost", "ready"),

    # OVERWHELMED
    ("Do not do to others what angers you if done to you by others.", "overwhelmed", "ready"),
    ("Contentment is natural wealth, luxury is artificial poverty.", "overwhelmed", "ready"),
    ("Prefer knowledge to wealth, for the one is transitory, the other perpetual.", "overwhelmed", "ready"),
    ("Be slow to fall into friendship, but when you are in, continue firm and constant.", "overwhelmed", "ready"),
    ("The wisest have the most authority.", "overwhelmed", "ready"),

    # MINDSET / GENERAL
    ("Education is the kindling of a flame, not the filling of a vessel.", "stuck", "ready"),
    ("By all means, marry. If you get a good wife, you'll be happy. If you get a bad one, you'll become a philosopher.", "lost", "ready"),
    ("The only real wisdom is knowing you know nothing.", "overwhelmed", "ready"),
    ("Once made equal to man, woman becomes his superior.", "lost", "ready"),
    ("By doing nothing we learn to do ill.", "lazy", "ready"),
    ("Prefer knowledge to wealth, for the one is transitory, the other perpetual.", "procrastinator", "ready"),
    ("The misuse of language induces evil in the soul.", "doomscroller", "ready"),
    ("Think not those faithful who praise all thy words and actions.", "quitter", "ready"),
    ("Our prayers should be for blessings in general, for God knows best what is good for us.", "overwhelmed", "ready"),
    ("The comic and the tragic lie inseparably close to each other.", "stuck", "ready"),
    ("When the debate is lost, slander becomes the tool of the loser.", "quitter", "ready"),
    ("No man has the right to be an amateur in the matter of physical training.", "lazy", "ready"),
    ("From the deepest desires often come the deadliest hate.", "doomscroller", "ready"),
    ("The soul takes nothing with her to the other world but her education and her culture.", "procrastinator", "ready"),
    ("Worthless people live only to eat and drink; people of worth eat and drink only to live.", "lazy", "ready"),
    ("Let him that would move the world, first move himself.", "stuck", "ready"),
    ("Beauty is a short-lived tyranny.", "doomscroller", "ready"),
    ("As for me, all I know is that I know nothing.", "quitter", "ready"),
    ("Falling is not the opposite of winning. Not trying is.", "lazy", "ready"),

    # Extra
    ("The first step toward wisdom is silence.", "stuck", "ready"),
    ("All men's souls are immortal, but the souls of the righteous are immortal and divine.", "lost", "ready"),
    ("The highest form of human excellence is to question oneself and others.", "procrastinator", "ready"),
    ("Think for yourself. Question everything.", "doomscroller", "ready"),
    ("I know what I do not know — and that is my advantage.", "quitter", "ready"),

    # ── PLATO'S APOLOGY ─────────────────────────────────────────────────────────
    ("A man who is good for anything ought not to calculate the chance of living or dying.", "quitter", "ready"),
    ("I am the gadfly which the god has given the state, and all day long I never cease to settle here and there.", "procrastinator", "ready"),
    ("I was really too honest a man to be a politician and live.", "overwhelmed", "ready"),
    ("The difficulty is not to avoid death, but to avoid unrighteousness.", "stuck", "ready"),
    ("I had rather go to death than to speak in a certain way.", "quitter", "ready"),
    ("No evil can happen to a good man, either in life or after death.", "stuck", "ready"),
    ("The unexamined life is not worth living.", "lost", "ready"),
    ("I do nothing but go about persuading you all, old and young alike, to care for your souls.", "doomscroller", "ready"),
    ("Are you not ashamed that you give your attention to acquiring as much money as possible, and to reputation and honor, and care so little about wisdom and truth?", "procrastinator", "ready"),
    ("I am not afraid of death. I just don't want to be there when it happens.", "overwhelmed", "ready"),
    ("Think not of the opinion of the many, but of the opinion of the one who knows about justice.", "stuck", "ready"),
    ("It is not the same thing to live and to live well.", "lazy", "ready"),
    ("Good men must not be anxious about life or death.", "quitter", "ready"),
    ("The hour of departure has arrived, and we go our ways.", "stuck", "ready"),
    ("Be of good cheer about death, and know this as a truth: that no evil can happen to a good man.", "lost", "ready"),
    ("I cannot teach anyone anything. I can only make them think.", "overwhelmed", "ready"),
    ("I know that I am intelligent because I know that I know nothing.", "stuck", "ready"),
    ("I am not going to alter my conduct for the sake of escaping death.", "quitter", "ready"),
    ("My death is a matter of little moment. It is the living that matters.", "stuck", "ready"),
    ("If I am an inventor, I can claim to have invented nothing useful.", "lazy", "ready"),
    ("I have never been a teacher of anyone.", "overwhelmed", "ready"),

    # ── PLATO'S CRITO ───────────────────────────────────────────────────────────
    ("Not life, but a good life, is to be chiefly valued.", "procrastinator", "ready"),
    ("One should not return an injustice or do any wrong to anyone, no matter what one may have suffered from them.", "lazy", "ready"),
    ("A man must take with him into the other world the full measure of his understanding.", "lost", "ready"),
    ("The opinions of the wise are more valuable than the opinions of the many.", "stuck", "ready"),
    ("Do not let your eagerness for victory overcome your sense of what is right.", "doomscroller", "ready"),
    ("A state, if well-governed, is like a ship that is steered by the wisdom of the captain.", "stuck", "ready"),
    ("The really important thing is not to live, but to live well.", "procrastinator", "ready"),
    ("We must follow the opinions of the one who understands.", "lost", "ready"),
    ("It is never right to do an injustice, even in return for an injustice.", "lazy", "ready"),
    ("The voice of the people is not always the voice of truth.", "doomscroller", "ready"),
    ("One who has knowledge will have the courage to act on it.", "quitter", "ready"),
    ("You cannot live well without knowledge of what is good.", "stuck", "ready"),

    # ── PLATO'S PHAEDO ──────────────────────────────────────────────────────────
    ("True philosophers are always occupied with the practice of dying.", "stuck", "ready"),
    ("The soul takes nothing with her to the other world but her education and her culture.", "procrastinator", "ready"),
    ("To be afraid of death is to think you are wise when you are not.", "quitter", "ready"),
    ("Wisdom is the purification of the soul from the body.", "overwhelmed", "ready"),
    ("The body is a source of endless trouble to us.", "lazy", "ready"),
    ("The soul is immortal and imperishable, and our souls will truly exist in the other world.", "lost", "ready"),
    ("He who has lived as a philosopher has no cause to fear death.", "stuck", "ready"),
    ("True wisdom comes to each of us when we realize how little we understand about life.", "lost", "ready"),
    ("The practice of philosophy is nothing but the practice of dying and being dead.", "quitter", "ready"),
    ("Every seeker of wisdom knows that philosophy is the study of death.", "stuck", "ready"),
    ("The soul which is pure at departing draws after her no bodily taint.", "overwhelmed", "ready"),
    ("Virtue is the harmony of the soul.", "lazy", "ready"),
    ("Death is the separation of the soul from the body.", "lost", "ready"),
    ("The pleasure of learning is the only pure pleasure.", "procrastinator", "ready"),
    ("Courage is knowing what not to fear.", "quitter", "ready"),

    # ── PLATO'S SYMPOSIUM ───────────────────────────────────────────────────────
    ("Love is the desire for the perpetual possession of the good.", "lazy", "ready"),
    ("The greatest wisdom is to be convinced of your own ignorance.", "stuck", "ready"),
    ("Every one of us is but a fragment of a complete whole.", "lost", "ready"),
    ("Love is the search for beauty and goodness, a journey of the soul.", "overwhelmed", "ready"),
    ("The beautiful is the object of love.", "doomscroller", "ready"),
    ("Love is born into every human being, calling us back to our original nature.", "lost", "ready"),
    ("Through love, we ascend from physical beauty to the beauty of the soul.", "doomscroller", "ready"),
    ("The lover goes beyond the beloved and seeks the divine.", "quitter", "ready"),
    ("The path of love leads from the particular to the universal.", "stuck", "ready"),
    ("In loving the beautiful, we love what is eternal.", "lost", "ready"),
    ("The greatest knowledge is to know the essence of love.", "overwhelmed", "ready"),

    # ── PLATO'S REPUBLIC ─────────────────────────────────────────────────────────
    ("The heaviest penalty for declining to rule is to be ruled by someone inferior.", "procrastinator", "ready"),
    ("The measure of a man is what he does with power.", "lazy", "ready"),
    ("Justice is the bond that holds society together.", "doomscroller", "ready"),
    ("Education is the art of turning the eye of the soul toward the light.", "stuck", "ready"),
    ("People must be just before they can be happy.", "quitter", "ready"),
    ("The greatest wealth is to live content with little.", "overwhelmed", "ready"),
    ("The direction in which education starts a man will determine his future life.", "procrastinator", "ready"),
    ("A good decision is based on knowledge, not on numbers.", "doomscroller", "ready"),
    ("The soul of man is immortal and imperishable.", "lost", "ready"),
    ("Our object in the construction of the state is the greatest happiness of the whole.", "stuck", "ready"),
    ("The price good men pay for indifference to public affairs is to be ruled by evil men.", "procrastinator", "ready"),
    ("Better to be poor and just than rich and unjust.", "lazy", "ready"),
    ("Courage is a kind of salvation, preserving our beliefs about what is and is not to be feared.", "quitter", "ready"),
    ("The end of wisdom is to desire what is good.", "lost", "ready"),
    ("The true quality of a ruler is to rule for the benefit of the ruled.", "overwhelmed", "ready"),
    ("The soul of the just man is like a well-tuned instrument.", "stuck", "ready"),
    ("The love of truth is the beginning of all wisdom.", "doomscroller", "ready"),
    ("The greatest punishment for doing wrong is to become like the wrongdoer.", "procrastinator", "ready"),
    ("Ignorance is the root of all evil.", "stuck", "ready"),
    ("The man who does not know the truth will be like a dreamer.", "lost", "ready"),
    ("The pursuit of truth is the only path to freedom.", "quitter", "ready"),

    # ── PLATO'S GORGIAS ─────────────────────────────────────────────────────────
    ("It is better to suffer wrong than to do wrong.", "lazy", "ready"),
    ("The true orator must be just and know what is right.", "doomscroller", "ready"),
    ("Rhetoric is the art of ruling the minds of men.", "procrastinator", "ready"),
    ("The good man will not lie, even in the face of death.", "quitter", "ready"),
    ("To do wrong is the greatest of evils.", "lost", "ready"),
    ("The just man is happy, the unjust man is miserable.", "stuck", "ready"),
    ("Power without wisdom is worthless.", "overwhelmed", "ready"),
    ("The wise man seeks to become better, not just to appear better.", "lazy", "ready"),
    ("No one who is truly good ever suffers harm.", "stuck", "ready"),
    ("The art of persuasion without wisdom is dangerous.", "doomscroller", "ready"),
    ("The true leader seeks the good of the people, not his own.", "overwhelmed", "ready"),
    ("The soul that is ordered and disciplined is the soul that is good.", "lost", "ready"),
    ("The disciplined man finds freedom.", "lazy", "ready"),
    ("A man must be his own master.", "procrastinator", "ready"),

    # ── PLATO'S MENO ─────────────────────────────────────────────────────────────
    ("The knowledge of good and evil is the foundation of all virtue.", "stuck", "ready"),
    ("What we call learning is only a process of recollection.", "lost", "ready"),
    ("Virtue is either wholly or partly wisdom.", "procrastinator", "ready"),
    ("There is no teaching, only reminding.", "stuck", "ready"),
    ("The soul is immortal and has learned everything.", "lost", "ready"),
    ("To know the good is to do the good.", "lazy", "ready"),
    ("The truth about reality is always in our souls.", "doomscroller", "ready"),
    ("Seeking and learning are nothing but recollection.", "overwhelmed", "ready"),
    ("Knowledge is the only thing that makes a man good.", "procrastinator", "ready"),
    ("The right opinion is as good as knowledge until it walks away.", "quitter", "ready"),
    ("All nature is akin, and the soul has learned everything.", "lost", "ready"),
    ("One thing I am certain of: that no one learns anything from me.", "stuck", "ready"),

    # ── XENOPHON'S MEMORABILIA ──────────────────────────────────────────────────
    ("The best men are those who spend their time trying to become better.", "procrastinator", "ready"),
    ("The good man is always happy, the bad man always miserable.", "stuck", "ready"),
    ("There is only one good: knowledge. There is only one evil: ignorance.", "doomscroller", "ready"),
    ("The wise man learns from the mistakes of others.", "lazy", "ready"),
    ("Friendship is the greatest of all human goods.", "lost", "ready"),
    ("A man should be as careful in choosing his friends as in choosing his path.", "doomscroller", "ready"),
    ("The hardest thing to teach is unlearning what we think we know.", "overwhelmed", "ready"),
    ("It is the mark of a good leader to care for those he leads.", "stuck", "ready"),
    ("No man is wise enough to govern without the help of others.", "lost", "ready"),
    ("If you want to be a good leader, you must first learn to be a good follower.", "procrastinator", "ready"),
    ("The greatest victory is the one over yourself.", "quitter", "ready"),
    ("The idle man wastes his life as if it were infinite.", "lazy", "ready"),
    ("Choose wisely, for your habits become your character.", "procrastinator", "ready"),
    ("A man without self-control is like a city without walls.", "lazy", "ready"),
    ("The wise man orders his life with reason and purpose.", "overwhelmed", "ready"),
    ("To know what is right and not do it is the worst cowardice.", "quitter", "ready"),
    ("The most useful knowledge is the knowledge of how to live well.", "stuck", "ready"),
    ("The man who makes everything that leads to happiness depends upon himself is wise.", "lost", "ready"),
    ("It is disgraceful to grow old through sheer carelessness before seeing what you are capable of becoming.", "quitter", "ready"),
    ("The willing man finds a way, the unwilling finds an excuse.", "procrastinator", "ready"),
    ("A life without examination is not worth examining.", "stuck", "ready"),
    ("Better to be healthy in body and mind than wealthy in gold.", "overwhelmed", "ready"),
    ("The noblest pursuit is the pursuit of wisdom.", "doomscroller", "ready"),
    ("Every man has the power to become better than he is.", "lazy", "ready"),
    ("A man's character is his fate.", "stuck", "ready"),
    ("The beginning of wisdom is the desire to learn.", "lost", "ready"),

    # ── EPICTETUS ────────────────────────────────────────────────────────────────
    ("It's not what happens to you, but how you react to it that matters.", "stuck", "ready"),
    ("The greater the difficulty, the more glory in surmounting it.", "quitter", "ready"),
    ("People are not disturbed by things, but by the view they take of them.", "overwhelmed", "ready"),
    ("Wealth consists not in having great possessions, but in having few wants.", "procrastinator", "ready"),
    ("No great thing is created suddenly.", "lazy", "ready"),
    ("First say to yourself what you would be, then do what you have to do.", "procrastinator", "ready"),
    ("He who laughs at himself never runs out of things to laugh at.", "stuck", "ready"),
    ("The key is to keep company only with people who uplift you.", "doomscroller", "ready"),
    ("If you want to improve, be content to be seen as foolish.", "lazy", "ready"),
    ("Make the best use of what is in your power, and take the rest as it happens.", "overwhelmed", "ready"),
    ("Freedom is the only worthy goal in life.", "quitter", "ready"),
    ("Don't explain your philosophy. Embody it.", "procrastinator", "ready"),
    ("Circumstances don't make the man. They reveal him.", "stuck", "ready"),
    ("The soul is like a bowl of water. When the wind passes over it, it ripples.", "doomscroller", "ready"),
    ("You are a little soul carrying around a corpse.", "lost", "ready"),
    ("Whoever would be free, let him wish for nothing that is not in his power.", "overwhelmed", "ready"),
    ("Progress is not achieved by luck or accident, but by working on yourself daily.", "lazy", "ready"),
    ("The happiness of your life depends upon the quality of your thoughts.", "doomscroller", "ready"),
    ("Some things are up to us, and some things are not.", "overwhelmed", "ready"),
    ("Don't be concerned with what others think. Be concerned with what is true.", "stuck", "ready"),
    ("When you do something noble, don't stop to count the cost.", "quitter", "ready"),
    ("The philosopher's lecture room is a hospital. You didn't leave in pleasure, but in pain.", "procrastinator", "ready"),
    ("If someone can refute me, show me I'm wrong, I'll gladly change.", "stuck", "ready"),
    ("Keep your attention focused entirely on what is truly your own concern.", "lazy", "ready"),
    ("You become what you give your attention to.", "doomscroller", "ready"),

    # ── SENECA ───────────────────────────────────────────────────────────────────
    ("It is not that we have a short time to live, but that we waste much of it.", "procrastinator", "ready"),
    ("While we are waiting for life, life passes.", "lazy", "ready"),
    ("Luck is what happens when preparation meets opportunity.", "quitter", "ready"),
    ("We suffer more in imagination than in reality.", "overwhelmed", "ready"),
    ("Sometimes even to live is an act of courage.", "stuck", "ready"),
    ("The whole future lies in uncertainty: live immediately.", "procrastinator", "ready"),
    ("It is the power of the mind to be unconquerable.", "quitter", "ready"),
    ("Associate with people who are likely to improve you.", "doomscroller", "ready"),
    ("Difficulties strengthen the mind, as labor does the body.", "lazy", "ready"),
    ("True happiness is to enjoy the present, without anxious dependence upon the future.", "overwhelmed", "ready"),
    ("If a man knows not to which port he sails, no wind is favorable.", "lost", "ready"),
    ("Life is long if you know how to use it.", "procrastinator", "ready"),
    ("Fire tests gold, suffering tests brave men.", "quitter", "ready"),
    ("The greatest wealth is a poverty of desires.", "overwhelmed", "ready"),
    ("It is not the man who has too little, but the man who craves more, that is poor.", "procrastinator", "ready"),
    ("Begin at once to live, and count each day as a separate life.", "lazy", "ready"),
    ("A happy life is one which is in accordance with its own nature.", "stuck", "ready"),
    ("No man was ever wise by chance.", "lost", "ready"),
    ("We should every night call ourselves to account.", "doomscroller", "ready"),
    ("The mind that is anxious about future events is miserable.", "overwhelmed", "ready"),
    ("He who is brave is free.", "quitter", "ready"),
    ("Most powerful is he who has himself in his own power.", "lazy", "ready"),
    ("Life without the courage for death is a slavery.", "stuck", "ready"),
    ("Things do not touch the soul. Our opinions are the only things that disturb us.", "doomscroller", "ready"),
    ("Wherever there is a human being, there is an opportunity for kindness.", "lost", "ready"),

    # ── MARCUS AURELIUS ──────────────────────────────────────────────────────────
    ("You have power over your mind, not outside events. Realize this and you will find strength.", "overwhelmed", "ready"),
    ("The happiness of your life depends upon the quality of your thoughts.", "doomscroller", "ready"),
    ("Waste no more time arguing about what a good man should be. Be one.", "procrastinator", "ready"),
    ("The best revenge is to be unlike him who performed the injury.", "lazy", "ready"),
    ("Very little is needed to make a happy life. It is all within yourself.", "stuck", "ready"),
    ("When you arise in the morning, think of what a privilege it is to be alive.", "lost", "ready"),
    ("The soul becomes dyed with the color of its thoughts.", "doomscroller", "ready"),
    ("If you are distressed by anything external, the pain is not due to the thing itself.", "overwhelmed", "ready"),
    ("The impediment to action advances action. What stands in the way becomes the way.", "quitter", "ready"),
    ("Accept the things to which fate binds you, and love the people with whom fate brings you together.", "stuck", "ready"),
    ("You could leave life right now. Let that determine what you do and say and think.", "procrastinator", "ready"),
    ("The universe is change. Our life is what our thoughts make it.", "lost", "ready"),
    ("It is not death that a man should fear, but he should fear never beginning to live.", "lazy", "ready"),
    ("Dwell on the beauty of life. Watch the stars, and see yourself running with them.", "stuck", "ready"),
    ("The soul must first be set free from passions to achieve tranquility.", "overwhelmed", "ready"),
    ("Nothing has such power to broaden the mind as the ability to investigate systematically.", "doomscroller", "ready"),
    ("The nearer a man comes to a calm mind, the closer he is to strength.", "quitter", "ready"),
    ("Loss is nothing else but change, and change is Nature's delight.", "stuck", "ready"),
    ("Look well into yourself. There is a source of strength which will always spring up.", "lost", "ready"),
    ("Do not act as if you were going to live a thousand years.", "procrastinator", "ready"),
    ("The universal beauty is like a river flowing, ever changing.", "doomscroller", "ready"),
    ("Let each thing you would do, say, or intend, be like that of a dying person.", "lazy", "ready"),
    ("Confine yourself to the present. The past and future are not in your power.", "overwhelmed", "ready"),
    ("Be tolerant with others and strict with yourself.", "quitter", "ready"),
    ("What we do now echoes in eternity.", "stuck", "ready"),

    # ── ORIGINAL SOCRATIC-STYLE ─────────────────────────────────────────────────
    ("A question unanswered is a cage you choose to live in.", "lost", "ready"),
    ("The opposite of wisdom is not ignorance, it is certainty.", "stuck", "ready"),
    ("You cannot think your way into a new life. But you can question your way out of the old one.", "procrastinator", "ready"),
    ("The mind that never questions is the mind that never grows.", "doomscroller", "ready"),
    ("What you avoid tells you more about yourself than what you pursue.", "lazy", "ready"),
    ("Every end is a beginning wearing a different mask.", "quitter", "ready"),
    ("The walls we build around ourselves are not made of stone, but of questions we refused to ask.", "stuck", "ready"),
    ("To scroll is to choose the thoughts of others over your own.", "doomscroller", "ready"),
    ("The only person who can hold you back is the one you see in the mirror.", "procrastinator", "ready"),
    ("Rest is not surrender. The bow that is always bent will break.", "overwhelmed", "ready"),
    ("The seed of greatness is buried in the soil of repetition.", "lazy", "ready"),
    ("You don't need a new beginning. You need to finish what you started.", "quitter", "ready"),
    ("The voice you avoid is the voice you most need to hear.", "lost", "ready"),
    ("A quiet mind is not an empty mind. It is a mind that has chosen its thoughts.", "overwhelmed", "ready"),
    ("The path to who you want to be is paved with who you are right now.", "stuck", "ready"),
    ("The noise you seek is the silence you fear.", "doomscroller", "ready"),
    ("Not every storm comes to disrupt your life. Some come to clear your path.", "quitter", "ready"),
    ("The question is not 'What should I do?' but 'Who should I become?'", "lost", "ready"),
    ("The heaviest burden is not the work you do, but the work you avoid.", "lazy", "ready"),
    ("Stillness is not the absence of noise. It is the presence of clarity.", "overwhelmed", "ready"),
    ("Courage is not the absence of fear. It is the decision that something else matters more.", "quitter", "ready"),
    ("The mind is a garden. What you water will grow.", "doomscroller", "ready"),
    ("You teach people how to treat you by what you tolerate.", "stuck", "ready"),
    ("The person you will be in five years is being built by the habits you start today.", "procrastinator", "ready"),
    ("There is no such thing as spare time. There is only life you choose to fill.", "lazy", "ready"),
    ("The wrong path taken consciously is wiser than the right path taken by accident.", "lost", "ready"),
    ("Your attention is the only currency that matters. Spend it wisely.", "doomscroller", "ready"),
    ("The obstacle is not the enemy. The refusal to move is.", "stuck", "ready"),
    ("You cannot control the waves, but you can learn to surf.", "overwhelmed", "ready"),
    ("The person who says it cannot be done should not interrupt the person doing it.", "procrastinator", "ready"),
    ("Letting go is not giving up. It is choosing what to carry.", "overwhelmed", "ready"),
    ("The only failure is the failure to begin.", "lazy", "ready"),
    ("What you seek is seeking you.", "lost", "ready"),
    ("The pain of discipline weighs ounces. The pain of regret weighs tons.", "quitter", "ready"),
    ("The world is full of voices. Wisdom is learning which one to follow.", "doomscroller", "ready"),
    ("A goal without a question is just a wish.", "stuck", "ready"),
    ("The bridge you need to cross only appears when you start walking.", "lost", "ready"),
    ("The smallest act of courage is worth more than the grandest intention.", "procrastinator", "ready"),
    ("Time is the only thief that returns nothing.", "lazy", "ready"),
    ("The mind is a mirror. Clean it daily.", "doomscroller", "ready"),
    ("Comparison is the thief of contentment.", "overwhelmed", "ready"),
    ("The fire that tests you is the fire that tempers you.", "quitter", "ready"),
    ("The questions we ask shape the answers we find.", "stuck", "ready"),
    ("Discipline is the bridge between goals and accomplishment.", "procrastinator", "ready"),
    ("The weight of the world is only as heavy as the strength you give it.", "overwhelmed", "ready"),
    ("A life lived on autopilot is a life not lived at all.", "doomscroller", "ready"),
    ("The longest journey begins with a single step away from yourself.", "lost", "ready"),
    ("You cannot pour from an empty vessel. Fill yourself first.", "overwhelmed", "ready"),
    ("The echo of your actions will outlive the memory of your excuses.", "lazy", "ready"),
    ("Every moment is a fresh beginning.", "stuck", "ready"),
    ("The silence between thoughts is where wisdom lives.", "doomscroller", "ready"),
    ("Do not wait for the perfect moment. Take the moment and make it perfect.", "procrastinator", "ready"),
    ("The cost of your comfort is the life you could be living.", "lazy", "ready"),
    ("When you change the way you look at things, the things you look at change.", "stuck", "ready"),
    ("Your past is a teacher, not a prison.", "lost", "ready"),
    ("The strength of a man is measured not by his muscles, but by his choices.", "quitter", "ready"),
    ("Kindness is the truest form of wisdom.", "lost", "ready"),
    ("The river of life flows whether you swim or stand still.", "lazy", "ready"),
    ("The cage you feel most is the one you cannot see.", "stuck", "ready"),
    ("The art of living is the art of knowing what to ignore.", "overwhelmed", "ready"),
    ("The roots of wisdom grow in the soil of humility.", "lost", "ready"),
    ("A restless mind is a restless life. Seek stillness.", "doomscroller", "ready"),
    ("The warrior fights not because he hates, but because he loves.", "quitter", "ready"),
    ("The value of a man is measured by what he gives, not by what he takes.", "procrastinator", "ready"),
    ("The seed of your greatest strength is planted in your deepest struggle.", "stuck", "ready"),
    ("Simplicity is the ultimate sophistication of the soul.", "overwhelmed", "ready"),
    ("Freedom is not doing what you want. It is wanting what is good.", "lost", "ready"),
    ("The shortcut is the longest road.", "procrastinator", "ready"),
    ("Your focus is your future. Watch where you point it.", "doomscroller", "ready"),
    ("The only way to make sense out of change is to plunge into it.", "quitter", "ready"),
    ("A single conversation across the kitchen table can change everything.", "stuck", "ready"),
    ("The door you keep walking past is the door you need to open.", "lazy", "ready"),
    ("What you resist persists. What you face fades.", "overwhelmed", "ready"),
    ("The quality of your questions determines the quality of your life.", "lost", "ready"),
    ("Do not confuse motion with progress.", "procrastinator", "ready"),
    ("The algorithm shows you what you already are. You choose what you become.", "doomscroller", "ready"),
]


def build_excel():
    wb = Workbook()

    # ── Sheet 1: Quotes Database ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Quotes"

    header_fill = PatternFill("solid", start_color="1A1A1A", end_color="1A1A1A")
    ready_fill = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
    edit_fill = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")

    thin = Border(
        left=Side(style="thin", color="C9A96E"),
        right=Side(style="thin", color="C9A96E"),
        top=Side(style="thin", color="C9A96E"),
        bottom=Side(style="thin", color="C9A96E"),
    )

    headers = ["#", "Quote", "Audience", "Caption A (Hook First)", "Caption B (Story First)", "Mood (AI fills this)", "Status", "Posted Date", "Post ID"]
    col_widths = [5, 60, 18, 80, 80, 30, 12, 16, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="C9A96E", name="Arial", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30

    audience_colors = {
        "procrastinator": "FFE5D0", "doomscroller": "D0E8FF",
        "stuck": "D0FFD6", "lazy": "FFF0D0",
        "quitter": "FFD0D0", "lost": "EDD0FF", "overwhelmed": "D0F0FF",
    }

    for i, (quote, audience, status) in enumerate(QUOTES, 1):
        row = i + 1
        caption_a = _build_caption(quote, audience, i - 1)
        caption_b = _build_caption(quote, audience, i)  # offset by 1 for variety
        aud_color = audience_colors.get(audience, "FFFFFF")
        row_fill = PatternFill("solid", start_color=aud_color, end_color=aud_color)

        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=2, value=quote).alignment = Alignment(wrap_text=True, vertical="top")
        aud_cell = ws.cell(row=row, column=3, value=audience)
        aud_cell.fill = row_fill
        aud_cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=4, value=caption_a).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5, value=caption_b).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=6, value="").alignment = Alignment(wrap_text=True, vertical="top")  # Mood column

        status_cell = ws.cell(row=row, column=7, value=status)
        status_cell.fill = ready_fill if status == "ready" else edit_fill
        status_cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.cell(row=row, column=8, value="").alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=9, value="").alignment = Alignment(horizontal="center", vertical="top")

        for col in range(1, 10):  # was 9, now 10 columns
            ws.cell(row=row, column=col).border = thin
            ws.cell(row=row, column=col).font = Font(name="Arial", size=10)

        ws.row_dimensions[row].height = 100

    ws.freeze_panes = "A2"

    # ── Sheet 2: Instructions ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 100

    instructions = [
        ("STOIC START — QUOTE DATABASE", True, "1A1A1A", "C9A96E"),
        ("", False, "FFFFFF", "000000"),
        ("HOW IT WORKS", True, "C9A96E", "1A1A1A"),
        ("1. Each quote has a story-driven caption (Hook → Story → Wisdom → CTA)", False, "F8F4EE", "1A1A1A"),
        ("2. Pipeline reads one row per day, picks by date, marks as posted.", False, "F8F4EE", "1A1A1A"),
        ("3. Claude AI only generates the image mood prompt — tiny API cost.", False, "F8F4EE", "1A1A1A"),
        ("4. To add more: just append rows below with the same column format.", False, "F8F4EE", "1A1A1A"),
        ("", False, "FFFFFF", "000000"),
        ("CAPTION STRUCTURE", True, "C9A96E", "1A1A1A"),
        ("📖 HOOK — Relatable scene that stops the scroll", False, "F8F4EE", "1A1A1A"),
        ("🔥 STORY — Personal moment of struggle or realisation", False, "F8F4EE", "1A1A1A"),
        ("💡 WISDOM — Socrates quote delivered as the breakthrough", False, "F8F4EE", "1A1A1A"),
        ("💬 CTA — Question that drives comments", False, "F8F4EE", "1A1A1A"),
        ("", False, "FFFFFF", "000000"),
        ("POSTING SCHEDULE", True, "C9A96E", "1A1A1A"),
        ("→ Pipeline runs 3x daily on GitHub Actions: 7:30, 13:00, 19:00 UTC", False, "F8F4EE", "1A1A1A"),
        ("→ Each run picks the next unposted quote by day-of-year", False, "F8F4EE", "1A1A1A"),
        ("→ Need 365 rows for a full year at 1/day; 1095 for 3/day", False, "F8F4EE", "1A1A1A"),
    ]

    for row_num, (text, bold, bg, fg) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=text)
        cell.font = Font(bold=bold, name="Arial", size=11, color=fg)
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws2.row_dimensions[row_num].height = 22

    # ── Sheet 3: Audience Guide ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Audience Guide")
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 50

    guide_headers = ["Audience", "Hook Style", "Story Angle"]
    for col, h in enumerate(guide_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="C9A96E", name="Arial")
        c.fill = PatternFill("solid", start_color="1A1A1A", end_color="1A1A1A")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    guide_data = [
        ("procrastinator", "Confrontational — calls out avoidance", "Personal story of waiting vs. finally starting"),
        ("doomscroller", "Mirror-holding — makes them see their behaviour", "Screen time regret, algorithm manipulation"),
        ("stuck", "Reframing — redefines stuck as opportunity", "Fear of change overcome by small steps"),
        ("lazy", "Punchy and energetic — short, direct", "Action-before-motivation realisation"),
        ("quitter", "Empathetic but firm", "Almost-quit story, the cost of giving up"),
        ("lost", "Philosophical and gentle", "Asking the right questions, self-discovery"),
        ("overwhelmed", "Calming and permission-giving", "Burnout story, learning to let go"),
    ]

    aud_colors_guide = {
        "procrastinator": "FFE5D0", "doomscroller": "D0E8FF", "stuck": "D0FFD6",
        "lazy": "FFF0D0", "quitter": "FFD0D0", "lost": "EDD0FF", "overwhelmed": "D0F0FF",
    }

    for row_num, (aud, hook, story) in enumerate(guide_data, 2):
        fill = PatternFill("solid", start_color=aud_colors_guide[aud], end_color=aud_colors_guide[aud])
        for col, val in enumerate([aud, hook, story], 1):
            c = ws3.cell(row=row_num, column=col, value=val)
            c.fill = fill
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
        ws3.row_dimensions[row_num].height = 40

    wb.save("quotes.xlsx")
    print(f"✅ Created quotes.xlsx with {len(QUOTES)} story-driven quotes across 3 sheets")


if __name__ == "__main__":
    build_excel()
