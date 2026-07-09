import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock
import sys
import json

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.content.quote_generator import _generate_quote_batch, generate_quotes, build_excel_from_quotes


@pytest.fixture
def sample_quotes():
    return [
        ("The only constant is change.", "stuck"),
        ("Knowledge is the beginning of wisdom.", "procrastinator"),
        ("Silence is the greatest teacher.", "overwhelmed"),
    ]


def test_generate_quote_batch_parses_json():
    """Should parse Claude JSON response into (quote, audience) pairs."""
    mock_response = MagicMock()
    mock_response.raise_for_status = MagicMock()
    mock_response.json.return_value = {
        "content": [{"text": json.dumps([
            {"quote": "Test quote one", "audience": "stuck"},
            {"quote": "Test quote two", "audience": "procrastinator"},
        ])}]
    }

    with patch("httpx.Client") as MockClient:
        mock_client = MagicMock()
        mock_client.__enter__.return_value.post.return_value = mock_response
        MockClient.return_value = mock_client

        results = _generate_quote_batch("fake_key", count=2)
        assert len(results) == 2
        assert results[0] == ("Test quote one", "stuck")
        assert results[1] == ("Test quote two", "procrastinator")


def test_generate_quotes_deduplicates():
    """Should deduplicate quotes by lowercase text."""
    with patch("src.content.quote_generator._generate_quote_batch") as mock_batch:
        mock_batch.return_value = [
            ("Duplicate quote", "stuck"),
            ("duplicate quote", "procrastinator"),  # same text, different case
            ("Unique quote", "lost"),
        ]
        results = generate_quotes("fake_key", target_count=3)
        assert len(results) == 2
        assert results[0] == ("Duplicate quote", "stuck")
        assert results[1] == ("Unique quote", "lost")


def test_generate_quotes_raises_on_empty():
    """Should raise RuntimeError if all generation attempts fail."""
    with patch("src.content.quote_generator._generate_quote_batch") as mock_batch:
        mock_batch.return_value = []
        with pytest.raises(RuntimeError, match="Failed to generate any quotes"):
            generate_quotes("fake_key", target_count=5)


def test_build_excel_from_quotes_creates_file(tmp_path, sample_quotes):
    """Should create a properly formatted Excel file."""
    output_path = tmp_path / "test_quotes.xlsx"
    result = build_excel_from_quotes(sample_quotes, output_path=str(output_path))

    assert Path(result).exists()
    assert output_path.exists()

    # Verify structure with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["Quotes"]

    # Check headers
    headers = [ws.cell(row=1, column=col).value for col in range(1, 10)]
    assert "#" in headers
    assert "Quote" in headers
    assert "Audience" in headers
    assert "Caption A (Hook First)" in headers
    assert "Caption B (Story First)" in headers
    assert "Status" in headers

    # Check data rows
    assert ws.cell(row=2, column=2).value == "The only constant is change."
    assert ws.cell(row=2, column=3).value == "stuck"
    assert ws.cell(row=2, column=7).value == "ready"

    # Check instructions sheet exists
    assert "Instructions" in wb.sheetnames
    assert "Audience Guide" in wb.sheetnames


def test_build_excel_captions_are_different(tmp_path):
    """Caption A and Caption B should be different for the same quote."""
    quotes = [("Test quote for captions.", "procrastinator")]
    output_path = tmp_path / "test_quotes.xlsx"
    build_excel_from_quotes(quotes, output_path=str(output_path))

    import openpyxl
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["Quotes"]

    caption_a = ws.cell(row=2, column=4).value
    caption_b = ws.cell(row=2, column=5).value
    assert caption_a != caption_b
    assert "Test quote for captions." in caption_a
    assert "Test quote for captions." in caption_b
