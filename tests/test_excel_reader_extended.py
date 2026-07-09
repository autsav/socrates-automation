import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.core.excel_reader import (
    AUDIENCE_TO_MOOD,
    VALID_MOODS,
    get_mood_prompt,
    mark_as_posted,
    read_todays_quote,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _create_quotes_excel(path: Path, rows: list[dict]):
    """Create a quotes.xlsx workbook with a Quotes sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotes"
    headers = ["#", "Quote", "Audience", "Caption", "Caption B", "Mood", "Status", "Posted Date", "Post ID"]
    ws.append(headers)
    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row.get("quote", ""),
            row.get("audience", ""),
            row.get("caption", ""),
            row.get("caption_b", ""),
            row.get("mood", ""),
            row.get("status", ""),
            row.get("posted_date", ""),
            row.get("post_id", ""),
        ])
    wb.save(path)


# ── read_todays_quote ────────────────────────────────────────────────────────

def test_read_todays_quote_file_not_found():
    """Should raise FileNotFoundError when the Excel file is missing."""
    with pytest.raises(FileNotFoundError):
        read_todays_quote("nonexistent_quotes.xlsx")


def test_read_todays_quote_no_ready_rows(tmp_path):
    """Should raise ValueError when all rows are skipped or already posted."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "Q1", "caption": "C1", "status": "skip"},
        {"quote": "Q2", "caption": "C2", "posted_date": "2024-01-01"},
    ])
    with pytest.raises(ValueError, match="No ready quotes left"):
        read_todays_quote(str(excel_path))


def test_read_todays_quote_empty_excel(tmp_path):
    """Should raise ValueError when the Excel file has no data rows."""
    excel_path = tmp_path / "quotes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotes"
    ws.append(["#", "Quote", "Audience", "Caption", "Caption B", "Mood", "Status", "Posted Date", "Post ID"])
    wb.save(excel_path)
    with pytest.raises(ValueError, match="No ready quotes left"):
        read_todays_quote(str(excel_path))


def test_read_todays_quote_happy_path(tmp_path):
    """Should return a ready quote with correct keys."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "Know thyself", "audience": "stuck", "caption": "Caption A"},
        {"quote": "Just do it", "audience": "procrastinator", "caption": "Caption B"},
    ])
    result = read_todays_quote(str(excel_path))
    assert isinstance(result, dict)
    assert "quote" in result
    assert "audience" in result
    assert "caption" in result
    assert "row_number" in result
    assert result["quote"] in ["Know thyself", "Just do it"]
    assert isinstance(result["status"], str)


def test_read_todays_quote_skips_missing_quote_or_caption(tmp_path):
    """Should skip rows with missing quote or caption."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "", "caption": "Only caption", "status": ""},
        {"quote": "Valid quote", "caption": "Valid caption", "status": ""},
    ])
    result = read_todays_quote(str(excel_path))
    assert result["quote"] == "Valid quote"


def test_read_todays_quote_skips_posted_rows(tmp_path):
    """Should skip rows that already have a posted date."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "Already posted", "caption": "Cap", "posted_date": "2024-01-01"},
        {"quote": "Fresh quote", "caption": "Cap", "posted_date": ""},
    ])
    result = read_todays_quote(str(excel_path))
    assert result["quote"] == "Fresh quote"


def test_read_todays_quote_picks_by_slot(tmp_path):
    """Should pick 3 unique quotes per day using slot-based formula (day × 3 + slot)."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "Quote 0", "caption": "Cap"},
        {"quote": "Quote 1", "caption": "Cap"},
        {"quote": "Quote 2", "caption": "Cap"},
    ])
    fake_dt = MagicMock()
    fake_dt.timetuple.return_value.tm_yday = 1  # day=1
    with patch("src.core.excel_reader.datetime") as mock_datetime:
        mock_datetime.now.return_value = fake_dt
        # Slot 0: (1*3+0) % 3 = 0 → Quote 0
        result0 = read_todays_quote(str(excel_path), slot=0)
        # Slot 1: (1*3+1) % 3 = 1 → Quote 1
        result1 = read_todays_quote(str(excel_path), slot=1)
        # Slot 2: (1*3+2) % 3 = 2 → Quote 2
        result2 = read_todays_quote(str(excel_path), slot=2)
    assert result0["quote"] == "Quote 0"
    assert result1["quote"] == "Quote 1"
    assert result2["quote"] == "Quote 2"
    # All three should be different
    assert len({result0["quote"], result1["quote"], result2["quote"]}) == 3


def test_read_todays_quote_audience_defaults_to_stuck(tmp_path):
    """Should default audience to 'stuck' when missing or empty."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "Quote", "caption": "Cap", "audience": ""},
    ])
    result = read_todays_quote(str(excel_path))
    assert result["audience"] == "stuck"


# ── get_mood_prompt ──────────────────────────────────────────────────────────

def _mock_httpx_response(status_code=200, text="dark_philosophical"):
    """Create a mock httpx response that mimics Anthropic API output."""
    mock = MagicMock()
    mock.status_code = status_code
    mock.raise_for_status = MagicMock()
    mock.json.return_value = {"content": [{"text": f"  {text}  "}]}
    return mock


def test_get_mood_prompt_happy_path():
    """Should return the mood returned by Claude."""
    mock_resp = _mock_httpx_response(text="dark_philosophical")
    with patch("httpx.Client") as MockClient:
        mock_client = MagicMock()
        mock_client.__enter__.return_value.post.return_value = mock_resp
        MockClient.return_value = mock_client

        mood = get_mood_prompt("A quote", "procrastinator", api_key="test-key")
        assert mood == "dark_philosophical"
        assert mock_client.__enter__.return_value.post.called


def test_get_mood_prompt_unknown_response_falls_back():
    """Should fallback to hardcoded map when Claude returns unrecognized text."""
    mock_resp = _mock_httpx_response(text="some_random_text")
    with patch("httpx.Client") as MockClient:
        mock_client = MagicMock()
        mock_client.__enter__.return_value.post.return_value = mock_resp
        MockClient.return_value = mock_client

        mood = get_mood_prompt("A quote", "doomscroller", api_key="test-key")
        assert mood == AUDIENCE_TO_MOOD["doomscroller"]


def test_get_mood_prompt_api_error_falls_back():
    """Should fallback to hardcoded mood when Claude API fails."""
    with patch("httpx.Client") as MockClient:
        mock_client = MagicMock()
        mock_client.__enter__.return_value.post.side_effect = Exception("API error")
        MockClient.return_value = mock_client

        mood = get_mood_prompt(
            "A quote", "overwhelmed",
            api_key="bad-key"
        )
        assert mood == AUDIENCE_TO_MOOD["overwhelmed"]


def test_get_mood_prompt_no_api_keys_hardcoded_fallback():
    """Should return hardcoded mood when no API key is provided."""
    mood = get_mood_prompt("A quote", "lazy", api_key="")
    assert mood == AUDIENCE_TO_MOOD["lazy"]


def test_get_mood_prompt_unknown_audience_defaults():
    """Should default to dark_philosophical for unknown audiences."""
    mock_resp = _mock_httpx_response(text="unknown_mood")
    with patch("httpx.Client") as MockClient:
        mock_client = MagicMock()
        mock_client.__enter__.return_value.post.return_value = mock_resp
        MockClient.return_value = mock_client

        mood = get_mood_prompt("A quote", "nonexistent_audience", api_key="test-key")
        assert mood == "dark_philosophical"


# ── mark_as_posted ───────────────────────────────────────────────────────────

def test_mark_as_posted_happy_path(tmp_path):
    """Should update Posted Date and Post ID for the matching row."""
    excel_path = tmp_path / "quotes.xlsx"
    _create_quotes_excel(excel_path, [
        {"quote": "Q1", "caption": "C1"},
        {"quote": "Q2", "caption": "C2"},
    ])
    mark_as_posted(str(excel_path), row_number=2, post_id="post_123")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Quotes"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][7] is None or rows[0][7] == ""   # row 1 unchanged
    assert rows[1][7] is not None                     # row 2 has posted date
    assert rows[1][8] == "post_123"                  # row 2 has post id


def test_mark_as_posted_file_not_found(tmp_path):
    """Should raise FileNotFoundError when the Excel file is missing."""
    excel_path = tmp_path / "missing.xlsx"
    with pytest.raises(FileNotFoundError):
        mark_as_posted(str(excel_path), row_number=1, post_id="post_123")


# ── Constants sanity checks ──────────────────────────────────────────────────

def test_valid_moods_contains_all_audience_values():
    """Every audience-to-mood value should be present in VALID_MOODS."""
    for mood in AUDIENCE_TO_MOOD.values():
        assert mood in VALID_MOODS


def test_audience_to_mood_has_expected_keys():
    """Sanity check that the fallback map has the expected audiences."""
    expected = {"procrastinator", "doomscroller", "stuck", "lazy", "quitter", "lost", "overwhelmed"}
    assert set(AUDIENCE_TO_MOOD.keys()) == expected
