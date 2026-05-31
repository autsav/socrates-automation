"""
Hybrid Quote Reader
- Reads quote + caption from Excel (FREE)
- Asks Claude only for image mood prompt (TINY API call ~£0.001)
- Falls back to hardcoded mood map if no API key
"""

import httpx
from datetime import datetime
from pathlib import Path
import openpyxl

from quote_generator import generate_fresh_quotes

REGEN_THRESHOLD = 10  # Trigger regeneration when fewer than this many quotes remain

# ── Fallback mood map (no Claude needed) ─────────────────────────────────────
AUDIENCE_TO_MOOD = {
    "procrastinator": "dark_philosophical",
    "doomscroller":   "dramatic_ancient",
    "stuck":          "cinematic_hopeful",
    "lazy":           "stark_minimal",
    "quitter":        "epic_warrior",
    "lost":           "mystical_greek",
    "overwhelmed":    "calm_stoic",
}

VALID_MOODS = list(AUDIENCE_TO_MOOD.values())


def _count_ready_quotes(ws) -> int:
    """Count how many rows are ready to post (not skipped, not posted)."""
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        status = row[6].value  # col G: Status
        posted = row[7].value  # col H: Posted Date
        quote = row[1].value   # col B: Quote
        caption = row[3].value # col D: Caption

        if not quote or not caption:
            continue
        if str(status).lower() == "skip":
            continue
        if posted:
            continue
        count += 1
    return count


def _current_slot() -> int:
    """Determine which daily slot we're in based on hour.
    Slot 0: morning (before 10 AM)    — 7:30 UTC
    Slot 1: afternoon (10-16)          — 13:00 UTC
    Slot 2: evening (after 16)         — 19:00 UTC
    """
    hour = datetime.now().hour
    if hour < 10:
        return 0
    elif hour < 16:
        return 1
    return 2


def read_todays_quote(
    excel_path: str = "quotes.xlsx",
    slot: int | None = None,
    api_key: str = "",
) -> dict:
    """
    Read today's quote from Excel using slot-based selection.
    Automatically regenerates fresh quotes if the pool is empty or low.
    Formula: (day_of_year × 3 + slot) % len(pool)
    Set slot=None for single-post-per-day fallback (original behavior).
    Returns dict: quote, audience, caption, row_number
    """
    path = Path(excel_path)

    # Auto-generate if file doesn't exist
    if not path.exists():
        if not api_key:
            raise FileNotFoundError(
                f"quotes.xlsx not found at {path.absolute()}\n"
                "Set ANTHROPIC_API_KEY to auto-generate fresh quotes."
            )
        print("  [excel] quotes.xlsx not found — generating fresh quotes...")
        generate_fresh_quotes(excel_path=str(path), api_key=api_key)

    wb = openpyxl.load_workbook(path)
    ws = wb["Quotes"]

    # Check remaining count and regenerate proactively if low
    remaining = _count_ready_quotes(ws)
    if remaining < REGEN_THRESHOLD:
        if api_key:
            print(f"  [excel] Only {remaining} quotes left — regenerating fresh batch...")
            generate_fresh_quotes(excel_path=str(path), api_key=api_key)
            # Reload workbook after regeneration
            wb = openpyxl.load_workbook(path)
            ws = wb["Quotes"]
        else:
            print(f"  [excel] Warning: only {remaining} quotes left. "
                  "Set ANTHROPIC_API_KEY to auto-regenerate.")

    # Collect all ready rows
    ready_rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_num   = row[0].value  # col A: #
        quote     = row[1].value  # col B: Quote
        audience  = row[2].value  # col C: Audience
        caption   = row[3].value  # col D: Caption
        caption_b = row[4].value  # col E: Caption Variant B
        status    = row[6].value  # col G: Status
        posted    = row[7].value  # col H: Posted Date

        if not quote or not caption:
            continue
        if str(status).lower() == "skip":
            continue
        if posted:  # already posted — skip
            continue

        ready_rows.append({
            "row_number": row_num,
            "quote":      str(quote).strip(),
            "audience":   str(audience).strip().lower() if audience else "stuck",
            "caption":    str(caption).strip(),
            "caption_b":  str(caption_b).strip() if caption_b else str(caption).strip(),
            "status":     str(status).strip(),
        })

    # If still empty after potential regeneration, raise error
    if not ready_rows:
        raise ValueError(
            "No ready quotes left in Excel!\n"
            "Set ANTHROPIC_API_KEY to auto-generate fresh quotes."
        )

    # Pick by slot-based formula: (day × 3 + slot) % pool_size
    day = datetime.now().timetuple().tm_yday
    if slot is None:
        slot = _current_slot()
    index = (day * 3 + slot) % len(ready_rows)
    chosen = ready_rows[index]
    return chosen


def get_mood_prompt(quote: str, audience: str, api_key: str = "") -> str:
    """
    Ask Claude Haiku for image mood — tiny prompt, minimal tokens.
    Falls back to hardcoded mood map on any error.
    """
    system = (
        "You pick ONE image mood for an Instagram quote post background. "
        f"Reply with ONLY one of these exact options: {', '.join(VALID_MOODS)}. "
        "Nothing else. No explanation. No punctuation."
    )
    user = f"Quote: {quote[:120]}\nAudience: {audience}"

    if api_key:
        try:
            transport = httpx.HTTPTransport(local_address="0.0.0.0")
            with httpx.Client(transport=transport) as client:
                resp = client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "Content-Type": "application/json",
                        "x-api-key": api_key,
                        "anthropic-version": "2023-06-01",
                    },
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 20,
                        "system": system,
                        "messages": [{"role": "user", "content": user}],
                    },
                    timeout=15,
                )
                resp.raise_for_status()
                raw = resp.json()["content"][0]["text"].strip().lower()

            for mood in VALID_MOODS:
                if mood in raw:
                    print(f"  [mood] Claude → {mood}")
                    return mood

            fallback = AUDIENCE_TO_MOOD.get(audience, "dark_philosophical")
            print(f"  [mood] Claude unknown '{raw}' → fallback {fallback}")
            return fallback

        except Exception as e:
            print(f"  [mood] Claude error ({e}) → fallback")

    # Hardcoded fallback
    fallback = AUDIENCE_TO_MOOD.get(audience, "dark_philosophical")
    print(f"  [mood] Fallback → {fallback} (no API)")
    return fallback


def mark_as_posted(excel_path: str, row_number: int, post_id: str):
    """
    Mark a row as posted in Excel with date + post ID.
    Called after successful Instagram post.
    """
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Quotes"]

    today = datetime.now().strftime("%Y-%m-%d")

    for row in ws.iter_rows(min_row=2):
        if row[0].value == row_number:
            row[7].value = today      # col H: Posted Date
            row[8].value = post_id    # col I: Post ID
            break

    wb.save(path)
    print(f"  [excel] Marked row {row_number} as posted ({today})")


if __name__ == "__main__":
    import os
    from dotenv import load_dotenv
    load_dotenv()

    # Test: read today's quote
    data = read_todays_quote("quotes.xlsx")
    print(f"Quote:    {data['quote'][:80]}...")
    print(f"Audience: {data['audience']}")
    print(f"Row:      {data['row_number']}")

    # Test: get mood (with fallback)
    mood = get_mood_prompt(
        quote=data["quote"],
        audience=data["audience"],
        api_key=os.getenv("ANTHROPIC_API_KEY", "")
    )
    print(f"Mood:     {mood}")
