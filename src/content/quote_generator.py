"""
Quote Generator — automatically regenerates fresh quotes when Excel is exhausted.
Uses Claude Haiku to generate new philosophical quotes, then builds captions
using the existing story-driven template system.
"""

from src.utils.logger import get_logger
logger = get_logger(__name__)

import json
import httpx
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reuse caption templates from src.content.generate_quotes_excel
from src.content.generate_quotes_excel import (
    HOOKS, STORIES, LESSONS, CTAS, _build_caption,
)

BATCH_SIZE = 90  # 1 month supply at 3 posts/day
API_URL = "https://api.anthropic.com/v1/messages"


def _generate_quote_batch(
    api_key: str,
    count: int = 45,
    temperature: float = 0.9,
) -> list[tuple[str, str]]:
    """
    Call Claude Haiku to generate a batch of (quote, audience) pairs.
    Uses direct HTTP (no SDK) to avoid IPv6 issues on macOS.
    """
    system = (
        "You are a master philosopher in the style of Socrates and ancient Greek Stoics. "
        "Generate profound, original philosophical quotes that feel timeless and thought-provoking. "
        "Each quote should be paired with a target audience that would most benefit from it. "
        "Return ONLY a valid JSON array. No markdown, no explanations, no preamble."
    )

    user = (
        f"Generate a JSON array of {count} original philosophical quotes. "
        "Each object must have exactly two fields:\n"
        '- "quote": a short, profound philosophical quote (1-2 sentences max)\n'
        '- "audience": one of [procrastinator, doomscroller, stuck, lazy, quitter, lost, overwhelmed]\n\n'
        "Guidelines:\n"
        "- Quotes must be original, not from any known philosopher\n"
        "- They should feel timeless, like ancient wisdom rediscovered\n"
        "- Each quote should resonate deeply with its target audience\n"
        "- Vary the tone: some confrontational, some gentle, some questioning\n\n"
        "Return ONLY valid JSON array. Example format:\n"
        '[{"quote": "The heaviest chains are the ones we forge from our own certainty.", "audience": "stuck"}]'
    )

    transport = httpx.HTTPTransport(local_address="0.0.0.0")
    with httpx.Client(transport=transport) as client:
        resp = client.post(
            API_URL,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 4096,
                "temperature": temperature,
                "system": system,
                "messages": [{"role": "user", "content": user}],
            },
            timeout=30,
        )
        resp.raise_for_status()
        raw = resp.json()["content"][0]["text"].strip()

    # Clean up markdown code fences if present
    if raw.startswith("```json"):
        raw = raw[7:]
    if raw.startswith("```"):
        raw = raw[3:]
    if raw.endswith("```"):
        raw = raw[:-3]
    raw = raw.strip()

    data = json.loads(raw)
    results = []
    for item in data:
        quote = item.get("quote", "").strip()
        audience = item.get("audience", "stuck").strip().lower()
        if quote and audience in HOOKS:
            results.append((quote, audience))

    return results


def generate_quotes(
    api_key: str,
    target_count: int = BATCH_SIZE,
) -> list[tuple[str, str]]:
    """
    Generate a full batch of fresh quotes using Claude Haiku.
    Makes multiple API calls if needed to reach target_count.
    """
    all_quotes = []
    attempts = 0
    max_attempts = 5

    logger.info(f"  [quotes] Generating {target_count} fresh quotes...")

    while len(all_quotes) < target_count and attempts < max_attempts:
        remaining = target_count - len(all_quotes)
        batch_size = min(remaining, 45)  # Claude Haiku output limit safety

        try:
            batch = _generate_quote_batch(api_key, count=batch_size)
            all_quotes.extend(batch)
            logger.info(f"  [quotes] Got {len(batch)} quotes (total: {len(all_quotes)}/{target_count})")
        except Exception as e:
            logger.info(f"  [quotes] Generation attempt {attempts + 1} failed: {e}")

        attempts += 1

    if not all_quotes:
        raise RuntimeError(
            "Failed to generate any quotes after multiple attempts. "
            "Check ANTHROPIC_API_KEY and network connection."
        )

    # Deduplicate by quote text (case-insensitive)
    seen = set()
    unique = []
    for quote, audience in all_quotes:
        key = quote.lower()
        if key not in seen:
            seen.add(key)
            unique.append((quote, audience))

    logger.info(f"  [quotes] Generated {len(unique)} unique quotes")
    return unique


def build_excel_from_quotes(
    quotes: list[tuple[str, str]],
    output_path: str = "quotes.xlsx",
) -> str:
    """
    Build a fresh quotes.xlsx from generated quote+audience pairs.
    Uses the exact same format as generate_quotes_excel.py.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"

    header_fill = PatternFill("solid", start_color="1A1A1A", end_color="1A1A1A")
    ready_fill = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
    edit_fill = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")

    thin = Border(
        left=Side(style="thin", color="C9A96E"),
        right=Side(style="thin", color="C9A96E"),
        top=Side(style="thin", color="C9A96E"),
        bottom=Side(style="thin", color="C9A96E"),
    )

    headers = [
        "#", "Quote", "Audience",
        "Caption A (Hook First)", "Caption B (Story First)",
        "Mood (AI fills this)", "Status", "Posted Date", "Post ID",
    ]
    col_widths = [5, 60, 18, 80, 80, 30, 12, 16, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="C9A96E", name="Arial", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30

    audience_colors = {
        "procrastinator": "FFE5D0", "doomscroller": "D0E8FF",
        "stuck": "D0FFD6", "lazy": "FFF0D0",
        "quitter": "FFD0D0", "lost": "EDD0FF", "overwhelmed": "D0F0FF",
    }

    for i, (quote, audience) in enumerate(quotes, 1):
        row = i + 1
        caption_a = _build_caption(quote, audience, i - 1)
        caption_b = _build_caption(quote, audience, i)
        aud_color = audience_colors.get(audience, "FFFFFF")
        row_fill = PatternFill("solid", start_color=aud_color, end_color=aud_color)

        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=2, value=quote).alignment = Alignment(wrap_text=True, vertical="top")
        aud_cell = ws.cell(row=row, column=3, value=audience)
        aud_cell.fill = row_fill
        aud_cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=4, value=caption_a).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5, value=caption_b).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=6, value="").alignment = Alignment(wrap_text=True, vertical="top")

        status_cell = ws.cell(row=row, column=7, value="ready")
        status_cell.fill = ready_fill
        status_cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.cell(row=row, column=8, value="").alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=9, value="").alignment = Alignment(horizontal="center", vertical="top")

        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin
            ws.cell(row=row, column=col).font = Font(name="Arial", size=10)

        ws.row_dimensions[row].height = 100

    ws.freeze_panes = "A2"

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 100
    instructions = [
        ("STOIC START — QUOTE DATABASE", True, "1A1A1A", "C9A96E"),
        ("", False, "FFFFFF", "000000"),
        ("HOW IT WORKS", True, "C9A96E", "1A1A1A"),
        ("1. Each quote has a story-driven caption (Hook → Story → Wisdom → CTA)", False, "F8F4EE", "1A1A1A"),
        ("2. Pipeline reads one row per day, picks by date, marks as posted.", False, "F8F4EE", "1A1A1A"),
        ("3. Claude AI generates both image mood AND fresh quotes when needed.", False, "F8F4EE", "1A1A1A"),
        ("4. Quotes auto-regenerate when the pool runs low — fully automated.", False, "F8F4EE", "1A1A1A"),
        ("", False, "FFFFFF", "000000"),
        ("CAPTION STRUCTURE", True, "C9A96E", "1A1A1A"),
        ("📖 HOOK — Relatable scene that stops the scroll", False, "F8F4EE", "1A1A1A"),
        ("🔥 STORY — Personal moment of struggle or realisation", False, "F8F4EE", "1A1A1A"),
        ("💡 WISDOM — Socrates quote delivered as the breakthrough", False, "F8F4EE", "1A1A1A"),
        ("💬 CTA — Question that drives comments", False, "F8F4EE", "1A1A1A"),
    ]
    for row_num, (text, bold, bg, fg) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=text)
        cell.font = Font(bold=bold, name="Arial", size=11, color=fg)
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws2.row_dimensions[row_num].height = 22

    # Audience Guide sheet
    ws3 = wb.create_sheet("Audience Guide")
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 50
    guide_headers = ["Audience", "Hook Style", "Story Angle"]
    for col, h in enumerate(guide_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="C9A96E", name="Arial")
        c.fill = PatternFill("solid", start_color="1A1A1A", end_color="1A1A1A")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    guide_data = [
        ("procrastinator", "Confrontational — calls out avoidance", "Personal story of waiting vs. finally starting"),
        ("doomscroller", "Mirror-holding — makes them see their behaviour", "Screen time regret, algorithm manipulation"),
        ("stuck", "Reframing — redefines stuck as opportunity", "Fear of change overcome by small steps"),
        ("lazy", "Punchy and energetic — short, direct", "Action-before-motivation realisation"),
        ("quitter", "Empathetic but firm", "Almost-quit story, the cost of giving up"),
        ("lost", "Philosophical and gentle", "Asking the right questions, self-discovery"),
        ("overwhelmed", "Calming and permission-giving", "Burnout story, learning to let go"),
    ]
    aud_colors_guide = {
        "procrastinator": "FFE5D0", "doomscroller": "D0E8FF", "stuck": "D0FFD6",
        "lazy": "FFF0D0", "quitter": "FFD0D0", "lost": "EDD0FF", "overwhelmed": "D0F0FF",
    }
    for row_num, (aud, hook, story) in enumerate(guide_data, 2):
        fill = PatternFill("solid", start_color=aud_colors_guide[aud], end_color=aud_colors_guide[aud])
        for col, val in enumerate([aud, hook, story], 1):
            c = ws3.cell(row=row_num, column=col, value=val)
            c.fill = fill
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
        ws3.row_dimensions[row_num].height = 40

    wb.save(output_path)
    logger.info(f"  [quotes] ✅ Created {output_path} with {len(quotes)} fresh quotes")
    return output_path


def generate_fresh_quotes(
    excel_path: str = "quotes.xlsx",
    api_key: str = "",
) -> str:
    """
    Main entry point: generate fresh quotes and rebuild the Excel file.
    Returns the path to the newly created Excel file.
    """
    if not api_key:
        raise ValueError(
            "ANTHROPIC_API_KEY is required to generate fresh quotes.\n"
            "Set it in your .env file or environment variables."
        )

    quotes = generate_quotes(api_key, target_count=BATCH_SIZE)
    return build_excel_from_quotes(quotes, output_path=excel_path)


if __name__ == "__main__":
    import os
    from dotenv import load_dotenv
    load_dotenv()

    key = os.getenv("ANTHROPIC_API_KEY", "")
    if key:
        generate_fresh_quotes(api_key=key)
    else:
        logger.info("Set ANTHROPIC_API_KEY in .env to generate quotes")
